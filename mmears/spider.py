import requests
from pprint import pprint
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'teachers-data'

status = {
    1: '申请',
    2: '拒绝',
    3: '签约完成',
    4: '面试',
    5: '培训',
    6: '第一轮试讲',
    7: '签约',
    8: '备用',
    9: '面试邀请1',
    10: '面试邀请2',
    11: '面试邀请3',
    12: '培训邀请1',
    13: '培训邀请2',
    14: '培训邀请3',
    15: '第一轮试讲邀请',
    16: '第二轮试讲邀请',
    17: '第二轮试讲',
    18: '弃权',
    19: '稍后面试',
    20: '试讲淘汰',
    21: '面试淘汰',
    22: '失联',
    23: '培训1v1邀请1',
    24: '培训1v1邀请2',
    25: '培训1v1邀请3',
    26: '过课件',
}

props = {
    'a': 'status',
    'b': 'name',
    'c': 'email',
    'd': 'skype',
    'e': 'teaching years',
    'f': 'citizenship',
    'g': 'channel',
    'h': 'start date',
    'i': 'countdown',
    'j': 'apply date',
    'k': 'comment'
}

for k, v in props.items():
    ws[k + '1'] = v

pi = 10000
pn = 1
baseUrl = 'https://s.mmears.com/api/teacher/apply/list?_pi={0}&_pn={1}'
url = baseUrl.format(pi, pn)

headers = {
    'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.96 Safari/537.36',
    'referer': 'https://s.mmears.com/teacher/apply/list',
    'cookie': 'userId=3081; token=7oLvwdL4WtT8L_ffulRdCSsQLasOsVLbyjHfcUtOAfKfNNvuKsvjsAAbq8zPMhW9AzVwb-cc1iNZQDdSLN2Z_qNKdx6Gvg9jTIuVPNSPpXL46NxMIvhcp3l0__fa73CJ; ph=ba2778f6; pm-Product=qxLWrJCIZJYBNz5Sy4B-A1in2t7WVO6YkXC5qwuMmD4tGvsQbHmIBfQthfJKxeSwQFGxWiIb27_1rHxYpT6de1iTKdaRRvKW9ZFXHE86Ul1mvP3sP06rP8vf5hHCxZanUxApCRNJfOiSGLlpL9vSScMEyAcwPQpogB-b2A_vC21_eQh65MVe6zyIyx2k1EKW'
}

while True:
    res = requests.get(url=url, headers=headers)
    if res.json()['code'] != 0 or len(res.json()['result']['data']) == 0:
        break

    datas = res.json()['result']['data']

    for i, data in enumerate(datas):
        ws['a' + str((pn - 1) * pi + (i + 2))] = status[data['status']]
        ws['b' + str((pn - 1) * pi + (i + 2))] = data['name']
        ws['c' + str((pn - 1) * pi + (i + 2))] = data['email']
        ws['d' + str((pn - 1) * pi + (i + 2))] = data['skype']
        ws['e' + str((pn - 1) * pi + (i + 2))] = data['teachAge']
        ws['f' + str((pn - 1) * pi + (i + 2))] = data['citizenship']
        ws['g' + str((pn - 1) * pi + (i + 2))] = data['hearFrom'] + '--' + data['hearFromDetail']
        ws['h' + str((pn - 1) * pi + (i + 2))] = data['startWorkDate']
        ws['i' + str((pn - 1) * pi + (i + 2))] = data['status']
        ws['j' + str((pn - 1) * pi + (i + 2))] = data['createTime']
        ws['k' + str((pn - 1) * pi + (i + 2))] = data['comment']

    print('page{0} is crawling down .... finished!'.format(pn))
    pn += 1
    url = baseUrl.format(pi, pn)


path = './waqiangjiao.xlsx'
wb.save(path)

